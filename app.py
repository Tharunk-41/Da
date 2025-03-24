import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io

def apply_summary_styles(ws_summary, inserted_rows, deleted_rows, change_counts):
    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center')
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    
    ws_summary.append(["Summary Report"])
    ws_summary['A1'].font = title_font
    ws_summary.append([])
    ws_summary.append(["Inserted Rows", inserted_rows])
    ws_summary.append(["Deleted Rows", deleted_rows])
    ws_summary['A3'].font = bold_font
    ws_summary['B3'].font = bold_font
    ws_summary['A4'].font = bold_font
    ws_summary['B4'].font = bold_font
    ws_summary['A3'].fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')  # Green for inserted
    ws_summary['A4'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # Red for deleted
    ws_summary.append([])
    
    ws_summary.append(["Column Name", "Changes Count"])
    ws_summary['A6'].fill = header_fill
    ws_summary['B6'].fill = header_fill
    ws_summary['A6'].font = header_font
    ws_summary['B6'].font = header_font
    ws_summary['A6'].alignment = center_align
    ws_summary['B6'].alignment = center_align
    ws_summary['A6'].border = thick_border
    ws_summary['B6'].border = thick_border
    
    row_idx = 7
    for col, count in change_counts.items():
        ws_summary.append([col, count])
        ws_summary[f'A{row_idx}'].alignment = center_align
        ws_summary[f'B{row_idx}'].alignment = center_align
        ws_summary[f'A{row_idx}'].border = thick_border
        ws_summary[f'B{row_idx}'].border = thick_border
        row_idx += 1
    for col in ['A', 'B']:
        ws_summary.column_dimensions[col].auto_size = True

def get_summary_counts(df_merged):
    inserted_rows = len(df_merged[df_merged['_merge'] == 'right_only'])
    deleted_rows = len(df_merged[df_merged['_merge'] == 'left_only'])
    change_counts = {}
    
    for col in df_merged.columns:
        if '_before' in col:
            base_col = col.replace('_before', '')
            valid_changes = (df_merged['_merge'] == 'both') & (df_merged[col] != df_merged[f'{base_col}_after'])
            count_changes = valid_changes.sum()
            if count_changes > 0:
                change_counts[base_col] = count_changes
    
    return inserted_rows, deleted_rows, change_counts

def create_summary_sheet(wb, df_merged):
    ws_summary = wb.create_sheet(title="Summary", index=0)
    inserted_rows, deleted_rows, change_counts = get_summary_counts(df_merged)
    apply_summary_styles(ws_summary, inserted_rows, deleted_rows, change_counts)

def compare_and_highlight(before_file, after_file):
    df_before, df_after = pd.read_csv(before_file), pd.read_csv(after_file)
    df_merged = merge_dataframes(df_before, df_after)
    
    with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        df_merged.to_excel(writer, sheet_name='Comparison', index=False)
    
    wb = load_workbook('output.xlsx')
    ws = wb['Comparison']
    create_summary_sheet(wb, df_merged)
    apply_highlighting(ws)
    
    ws.delete_cols(ws.max_column)
    wb.save('output.xlsx')
    wb.close()
    
    with open('output.xlsx', 'rb') as f:
        return io.BytesIO(f.read())

def merge_dataframes(df_before, df_after):
    df_merged = pd.merge(df_before, df_after, on='id', how='outer', suffixes=('_before', '_after'), indicator=True)
    column_order = ['id']
    for col in df_before.columns:
        if col != 'id':
            column_order.append(f'{col}_before')
            column_order.append(f'{col}_after')
    column_order.append('_merge')
    return df_merged[column_order]

def apply_highlighting(ws):
    red_fill, green_fill, yellow_fill = PatternFill(start_color='FF9999', fill_type='solid'), PatternFill(start_color='99FF99', fill_type='solid'), PatternFill(start_color='FFFF99', fill_type='solid')
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_status = row[-1].value
        if row_status == 'left_only':
            for cell in row: cell.fill = red_fill
        elif row_status == 'right_only':
            for cell in row: cell.fill = green_fill
        else:
            for i in range(1, len(row) - 1, 2):
                if row[i].value != row[i + 1].value:
                    row[i].fill = row[i + 1].fill = yellow_fill

st.set_page_config(page_title="CSV Comparison Tool", page_icon="📊", layout="centered")
st.title("🔍 CSV Comparison Tool")
st.markdown("Compare two CSV files and highlight differences.")

st.sidebar.header("Upload CSV Files")
before_file = st.sidebar.file_uploader("📂 Upload Before CSV", type=["csv"], key="before")
after_file = st.sidebar.file_uploader("📂 Upload After CSV", type=["csv"], key="after")

st.markdown("---")
tab1, tab2 = st.tabs(["📄 Instructions", "⚙️ Compare & Download"])

with tab1:
    st.subheader("How It Works")
    st.write("1. Upload the 'Before' and 'After' CSV files.")
    st.write("2. Click 'Compare' to generate a comparison report.")
    st.write("3. Download the Excel file with highlighted differences and a summary.")
    st.info("Red: Deleted rows | Green: New rows | Yellow: Modified values")

with tab2:
    if before_file and after_file:
        if st.button("🔍 Compare & Generate Report", use_container_width=True):
            with st.spinner("Processing... Please wait."):
                output_file = compare_and_highlight(before_file, after_file)
                st.success("✅ Comparison Complete! Click below to download the output file.")
                st.download_button("📥 Download Output.xlsx", output_file, file_name="output.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        st.warning("⚠️ Please upload both CSV files to proceed.")
