import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import os

class FileHandler:
    def __init__(self):
        self.delimiter_storage = {}
        self.readers = {
            "csv": pd.read_csv,
            "xlsx": pd.read_excel,
            "xls": pd.read_excel,
            "txt": self.read_txt  # Handle txt separately due to delimiter
        }

    def get_delimiter(self, file_type):
        label = f"Enter delimiter for {file_type} file (e.g., ',' or '\\t'):"
        delimiter = st.sidebar.text_input(label, key=f"delimiter_{file_type}")
        self.delimiter_storage[file_type] = delimiter if delimiter else None
        return self.delimiter_storage[file_type]

    def read_txt(self, file, file_type):
        delimiter = self.get_delimiter(file_type)
        if not delimiter:
            st.error(f"Delimiter is required for {file_type} files. Please enter one.")
            return None
        try:
            df = pd.read_csv(file, delimiter=delimiter)
            if df.shape[1] == 1:
                st.error(f"Delimiter '{delimiter}' seems incorrect for {file_type}.")
                return None
            return df
        except pd.errors.EmptyDataError:
            st.error("The file contains no data or has incorrect formatting.")
        except pd.errors.ParserError:
            st.error(f"Incorrect delimiter '{delimiter}' used. Please check the file format.")
        return None

    def read_file(self, file, file_type=None):
        file_extension = file.name.split(".")[-1].lower()
        file.seek(0)
        if file_extension == "txt" and file_type:
            return self.read_txt(file, file_type)
        elif file_extension in self.readers:
            return self.readers[file_extension](file)
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
    
    def add_reader(self, extension, reader_function):
        self.readers[extension] = reader_function


file_handler = FileHandler()

def highlight_headers(ws,key_columns):
    key_column_indices = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value in key_columns}
    header_fill = PatternFill(start_color='FFA500', end_color='4472C4', fill_type='solid')
    for i in key_column_indices.values():
        ws.cell(row=1, column=i).fill = header_fill

def create_summary_sheet(wb, df_merged, before_length, after_length, duplicates_before_count, duplicates_after_count,cross_duplicates_before_count, cross_duplicates_after_count):
    ws_summary = wb.create_sheet('Summary', index=0)
    # Calculate Inserted, Deleted Rows, and Change Counts
    inserted_rows = (df_merged['_merge'] == 'right_only').sum()
    deleted_rows = (df_merged['_merge'] == 'left_only').sum()
    change_counts = {
        col.replace('_before', ''): ((df_merged['_merge'] == 'both') & (df_merged[col].notnull()) & 
                                     (df_merged[f'{col.replace("_before", "_after")}'].notnull()) & 
                                     (df_merged[col] != df_merged[f'{col.replace("_before", "_after")}'])).sum()
        for col in df_merged.columns if '_before' in col
    }
    # Apply Summary Styles
    apply_summary_styles(ws_summary, inserted_rows, deleted_rows, change_counts, before_length, after_length, duplicates_before_count, duplicates_after_count,cross_duplicates_before_count, cross_duplicates_after_count)

def apply_summary_styles(ws_summary, inserted_rows, deleted_rows, change_counts, before_length, after_length, duplicates_before_count, duplicates_after_count,cross_duplicates_before_count, cross_duplicates_after_count):
    # Define Styles
    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    highlight_green = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
    highlight_red = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Title
    ws_summary.append(['Summary Report'])
    ws_summary['A1'].font = title_font
    ws_summary.append([])
    # Metrics Table
    ws_summary.append(['Metrics', 'Before', 'After'])
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}3'].font = header_font
        ws_summary[f'{col}3'].fill = header_fill
        ws_summary[f'{col}3'].alignment = Alignment(horizontal='center')
        ws_summary[f'{col}3'].border = thick_border
    data_rows = [
        ["Number of rows", before_length, after_length],
        ["Number of Duplicate IDs", duplicates_before_count, duplicates_after_count],[],
        ["Matching Keys in Both Files", before_length - (deleted_rows + duplicates_before_count+cross_duplicates_before_count)],
        ["Inserted Rows", inserted_rows],
        ["Deleted Rows", deleted_rows]
    ]
    for row in data_rows:
        ws_summary.append(row)
    ws_summary['A8'].fill = highlight_green
    ws_summary['A9'].fill = highlight_red
    # Format Metrics Table
    for row in range(4, 6):
        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].font = bold_font
            ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
            ws_summary[f'{col}{row}'].border = thick_border
    for row in range(7, 10):
        ws_summary[f'A{row}'].font = bold_font
        ws_summary[f'B{row}'].font = bold_font
    # Changes Count Section
    ws_summary.append([])
    ws_summary.append(['Column Name', 'Changes Count'])
    for col in ['A', 'B']:
        ws_summary[f'{col}11'].font = header_font
        ws_summary[f'{col}11'].fill = header_fill
        ws_summary[f'{col}11'].alignment = Alignment(horizontal='center')
        ws_summary[f'{col}11'].border = thick_border
    row_idx = 12
    for col, count in change_counts.items():
        ws_summary.append([col, count])
        ws_summary[f'A{row_idx}'].border = thick_border
        ws_summary[f'B{row_idx}'].border = thick_border
        ws_summary[f'A{row_idx}'].alignment = Alignment(horizontal='center')
        ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal='center')
        row_idx += 1
    # Auto-size Columns
    for col in ['A', 'B', 'C']:
        ws_summary.column_dimensions[col].auto_size = True


def identify_and_remove_duplicates(df, key_columns):
    duplicates = df[df.duplicated(subset=key_columns, keep=False)]
    df_no_duplicates = df[~df.duplicated(subset=key_columns, keep=False)]

    return  duplicates, df_no_duplicates

def identify_and_remove_cross_duplicates(df_before, df_after, key_columns):
    # Count occurrences of each key in both dataframes
    before_counts = df_before.groupby(key_columns, dropna=False).size().reset_index(name='count_before')
    after_counts = df_after.groupby(key_columns, dropna=False).size().reset_index(name='count_after')
    # Merge counts on key_columns
    counts = pd.merge(before_counts, after_counts, on=key_columns, how='outer').fillna(0).astype({'count_before': 'int', 'count_after': 'int'})
    # Identify cross-duplicate keys where count is imbalanced
    cross_duplicate_keys_df = counts.query(
        "(count_before == 1 and count_after > 1) or (count_after == 1 and count_before > 1)"
    )[key_columns]
    # Identify cross duplicate rows
    cross_duplicates_before = df_before.merge(cross_duplicate_keys_df, on=key_columns, how='inner')
    cross_duplicates_after = df_after.merge(cross_duplicate_keys_df, on=key_columns, how='inner')
    # Remove cross duplicates efficiently using merge with indicator
    df_before_no_cross_duplicates = df_before.merge(cross_duplicate_keys_df, on=key_columns, how='left', indicator=True)
    df_before_no_cross_duplicates = df_before_no_cross_duplicates.query("_merge == 'left_only'").drop(columns=['_merge'])
    df_after_no_cross_duplicates = df_after.merge(cross_duplicate_keys_df, on=key_columns, how='left', indicator=True)
    df_after_no_cross_duplicates = df_after_no_cross_duplicates.query("_merge == 'left_only'").drop(columns=['_merge'])

    return df_before_no_cross_duplicates, df_after_no_cross_duplicates, cross_duplicates_before, cross_duplicates_after

def create_duplicates_sheet(wb, duplicates_df, sheet_name):
    if not duplicates_df.empty:
        ws_duplicates = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(duplicates_df, index=False, header=True):
            ws_duplicates.append(r)
        highlight_headers(ws_duplicates, key_columns)

def merge_dataframes(df_before, df_after, key_columns):
    df_merged = pd.merge(df_before, df_after, on=key_columns, how='outer', suffixes=('_before', '_after'), indicator=True)
    column_order = key_columns.copy()
    for col in df_before.columns:
        if col not in key_columns:
            column_order.append(f'{col}_before')
            column_order.append(f'{col}_after')
    column_order.append('_merge')
    
    return df_merged[column_order].sort_values(key_columns).reset_index(drop=True)

def apply_highlighting(ws):
    red_fill,green_fill, yellow_fill = PatternFill(start_color='FF9999', fill_type='solid'), PatternFill(start_color='99FF99', fill_type='solid'), PatternFill(start_color='FFFF00', fill_type='solid')
    highlight_headers(ws, key_columns)
    key_column_indices = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value in key_columns}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_status=row[-1].value
        if row_status == 'left_only':
            for cell in row:
                cell.fill = red_fill
        elif row_status == 'right_only':
            for cell in row:
                cell.fill = green_fill
        else:
            for i in range(1,len(row)-1,2):
                if row[i].value != row[i+1].value:
                    if row[i].column not in key_column_indices.values():
                        row[i].fill = yellow_fill
                    if row[i+1].column not in key_column_indices.values():
                        row[i+1].fill = yellow_fill

def write_cross_duplicates_to_excel(writer, cross_duplicates_before, cross_duplicates_after, key_columns):
    workbook = writer.book
    # Write 'Cross_Duplicates_Before' only if not empty
    if not cross_duplicates_before.empty:
        cross_duplicates_before.to_excel(writer, sheet_name='Cross_Duplicates_Before', index=False)
        ws_before = workbook['Cross_Duplicates_Before']
        highlight_headers(ws_before, key_columns)
    # Write 'Cross_Duplicates_After' only if not empty
    if not cross_duplicates_after.empty:
        cross_duplicates_after.to_excel(writer, sheet_name='Cross_Duplicates_After', index=False)
        ws_after = workbook['Cross_Duplicates_After']
        highlight_headers(ws_after, key_columns)

def compare_and_highlight(df_before, df_after, key_columns):
    # Remove cross-duplicates
    df_before_a, df_after_a, cross_duplicates_before, cross_duplicates_after = identify_and_remove_cross_duplicates(df_before, df_after, key_columns)
    # Remove duplicates within each dataset
    duplicates_before, df_before_b= identify_and_remove_duplicates(df_before_a, key_columns)
    duplicates_after, df_after_b = identify_and_remove_duplicates(df_after_a, key_columns)
    # Merge DataFrames
    df_merged = merge_dataframes(df_before_b, df_after_b, key_columns)
    output_stream = io.BytesIO()
    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        df_merged.to_excel(writer, sheet_name='Comparison', index=False)
        ws = writer.sheets["Comparison"]
        apply_highlighting(ws)
        ws.delete_cols(ws.max_column)  
        create_summary_sheet(writer.book, df_merged, len(df_before), len(df_after), len(duplicates_before), len(duplicates_after),len(cross_duplicates_before), len(cross_duplicates_after))
        create_duplicates_sheet(writer.book, duplicates_before, sheet_name='Duplicates_Before')
        create_duplicates_sheet(writer.book, duplicates_after, sheet_name='Duplicates_After')
        write_cross_duplicates_to_excel(writer, cross_duplicates_before, cross_duplicates_after, key_columns)
    # Ensure everything is written properly
    output_stream.seek(0)
    return output_stream    

# StreamLit UI setup
st.set_page_config(
    page_title="CSV Comparison Tool",
    page_icon="📊",
    layout="centered",
)

st.title("CSV Comparison Tool")
st.markdown("Compare two CSV files and highlight differences.")

st.sidebar.header("Upload Files")
before_file = st.sidebar.file_uploader("Upload Before CSV", type=["csv", "xlsx", "xls", "txt"], key="before")
after_file = st.sidebar.file_uploader("Upload After CSV", type=["csv", "xlsx", "xls", "txt"], key="after")

# Ensure files are uploaded
if before_file and after_file:
    while True:
        df_before = file_handler.read_file(before_file,"before")  
        df_after = file_handler.read_file(after_file,"after")  

        if df_before is not None and df_after is not None:
            break  # Exit loop when both DataFrames are successfully read

        if df_before is None or df_after is None:
            st.error("One of the files could not be read. Please check the delimiter and file format.")
            st.stop()  # Stops execution to prevent further errors

    # Proceed after successful file reading
    common_columns = list(set(df_before.columns) & set(df_after.columns))
    if common_columns:
        key_columns = st.sidebar.multiselect("Select columns to merge on", common_columns)

    

if before_file:
    before_base_name = os.path.splitext(before_file.name)[0].replace('_before', '')
    output_file_name = f"{before_base_name}_Comparison.xlsx"

st.markdown("---")
tab1, tab2 = st.tabs(["Instructions", "Compare & Download"])

with tab1:
    st.subheader("How It Works:")
    st.write("1. Upload the 'Before' and 'After' CSV Files.")
    st.write("2. Select one or more key columns for comparison.")
    st.write("3. Click 'Compare' to generate a comparison report.")
    st.write("4. Download the Excel file with highlighted differences and a summary.")
    st.info("Red: Deleted rows | Green: New rows | Yellow: Modified values")

with tab2:
    if before_file and after_file and key_columns:
        if st.button("🔍 Generate Comparison Report", use_container_width=True):
            with st.spinner("Processing... Please wait."):
                output_file = compare_and_highlight(df_before, df_after, key_columns)
                st.success("Comparison Complete! Click below to download the output file.")
                st.download_button(f"Download {before_base_name}_Output File",output_file,file_name=output_file_name,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    else:
        st.warning("⚠️ Please upload both CSV files and select key columns to proceed.")
